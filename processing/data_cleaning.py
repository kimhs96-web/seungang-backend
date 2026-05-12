"""
데이터 정제 및 통계 추출 모듈
v21: 전년 동기 비교 분석 + 반복 장애 예방점검 기능 추가
"""
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import datetime

AC_DELETE_WORDS = [
    '이물질','조명','데마','콤','전도','끼임','누수','정전','청소','부주의',
    '충격','화재','멀티포스트','스피커','승차감','버튼','비고장','일정수립'
]
KEEP_COLS_IDX = [1, 3, 4, 7, 8, 9, 16, 23, 24, 28, 29]
FINAL_COLS = ['순번','발생일자','시간','장애유형','호선','역사','옥내/옥외',
              '신고내역','완료일자','완료시간','총복구시간','조치내역','장애종류']
WIDTH_MAP = {
    '순번':4.625,'발생일자':10.125,'시간':6.0,'장애유형':14.125,
    '호선':7.375,'역사':8.75,'옥내/옥외':10,'신고내역':40,
    '완료일자':10.125,'완료시간':7.375,'총복구시간':8.875,'조치내역':40,'장애종류':14
}
EL_PRIORITY = ['안전스위치','제어장치','속도조정장치','전장품','층감지장치',
               '동력장치','동력전달장치','제동장치','도어','부속장치']
ES_PRIORITY = ['안전스위치','동력전달장치','제동장치','제어장치','속도조정장치',
               '전장품','동력장치','스텝콤','핸드레일','멀티포스트','부속장치']
REPEAT_THRESHOLD = 3   # 분기 내 동일 역사+장애종류 N회 이상 → 반복장애


def normalize(s):
    if s is None: return ''
    return str(s).replace(' ','').replace('\u3000','').replace('\xa0','').strip()


def run_data_cleaning(src_path, cls_path, out_path):
    wb_src = load_workbook(str(src_path))
    wb_cls = load_workbook(str(cls_path))
    if '장애신고' in wb_src.sheetnames:
        ws_src = wb_src['장애신고']
    elif 'sheet1' in [s.lower() for s in wb_src.sheetnames]:
        matched = [s for s in wb_src.sheetnames if s.lower() == 'sheet1']
        ws_src = wb_src[matched[0]]
    else:
        ws_src = wb_src[wb_src.sheetnames[0]]

    col_map = {}
    for ci in range(1, ws_src.max_column + 1):
        v = ws_src.cell(2, ci).value
        if v: col_map[str(v).strip()] = ci

    COL = {
        'G':  col_map.get('장애유형', 7),
        'H':  col_map.get('호선', 8),
        'J':  col_map.get('관리번호', 10),
        'AC': col_map.get('조치내역', 29),
    }

    def should_delete(row):
        if '기타장애' in normalize(row[COL['G']-1].value) or \
           '인적장애' in normalize(row[COL['G']-1].value): return True
        ac = normalize(row[COL['AC']-1].value)
        for w in AC_DELETE_WORDS:
            if normalize(w) in ac: return True
        if '소음' in ac and '장력조정' in ac: return True
        return False

    all_rows = []
    for r in range(3, ws_src.max_row + 1):
        row = list(ws_src[r])
        if all(c.value is None for c in row): continue
        if not should_delete(row): all_rows.append(row)

    el_rows, es_rows = [], []
    for row in all_rows:
        mn = normalize(row[COL['J']-1].value)
        if 'EL' in mn: el_rows.append(row)
        elif 'ES' in mn: es_rows.append(row)

    def split_hoseon(rows):
        l1, l2 = [], []
        for row in rows:
            h = normalize(row[COL['H']-1].value)
            if '1호선' in h: l1.append(row)
            elif '2호선' in h: l2.append(row)
        return l1, l2

    el1, el2 = split_hoseon(el_rows)
    es1, es2 = split_hoseon(es_rows)

    def build_io_lookup(ws):
        d = {}
        for r in range(2, ws.max_row + 1):
            row = ws[r]
            for base in [0, 4]:
                if row[base].value:
                    try: hogi = str(int(row[base+1].value)) if row[base+1].value else ''
                    except: hogi = str(row[base+1].value) if row[base+1].value else ''
                    d[(str(row[base].value).strip(), hogi)] = row[base+2].value
        return d

    el_io = build_io_lookup(wb_cls['EL구분'])
    es_io = build_io_lookup(wb_cls['ES구분'])

    def get_io(lookup, yeoksa, gwanri_no):
        mn = str(gwanri_no) if gwanri_no else ''
        parts = mn.split('-')
        try: hogi = str(int(parts[-1])) if len(parts) >= 3 else ''
        except: hogi = parts[-1] if len(parts) >= 3 else ''
        return lookup.get((str(yeoksa).strip() if yeoksa else '', hogi), '')

    def build_fault_dict(ws):
        headers = [c.value for c in ws[1]]
        fd = {h: [] for h in headers if h}
        for r in range(2, ws.max_row + 1):
            for ci, cell in enumerate(ws[r]):
                if cell.value and ci < len(headers) and headers[ci]:
                    fd[headers[ci]].append(normalize(str(cell.value)))
        return fd

    el_fd = build_fault_dict(wb_cls['EL장애'])
    es_fd = build_fault_dict(wb_cls['ES장애'])

    def get_fault_type(fd, priority, ac_val):
        ac = normalize(ac_val)
        matched = []
        for cat, kws in fd.items():
            for kw in kws:
                if kw and kw in ac: matched.append(cat); break
        if not matched: return None
        matched.sort(key=lambda c: priority.index(c) if c in priority else 999)
        return matched[0]

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    def build_sheet(sn, src_rows, io_lookup, fd, priority):
        ws = wb_out.create_sheet(sn)
        for ci, name in enumerate(FINAL_COLS, 1):
            cell = ws.cell(1, ci, value=name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        seq = 1
        for row_src in src_rows:
            vals = [row_src[ci-1].value for ci in KEEP_COLS_IDX]
            yeoksa, gwanri_no, ac_val = vals[5], row_src[COL['J']-1].value, vals[10]
            io_val = get_io(io_lookup, yeoksa, gwanri_no)
            ft = get_fault_type(fd, priority, ac_val)
            if ft is None: continue
            out = [seq, vals[1], vals[2], vals[3], vals[4], vals[5], io_val,
                   vals[6], vals[7], vals[8], vals[9], vals[10], ft]
            for ci, v in enumerate(out, 1):
                cell = ws.cell(seq+1, ci, value=v)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            seq += 1
        for ci, name in enumerate(FINAL_COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = WIDTH_MAP.get(name, 10)

    build_sheet('1호선EL', el1, el_io, el_fd, EL_PRIORITY)
    build_sheet('2호선EL', el2, el_io, el_fd, EL_PRIORITY)
    build_sheet('1호선ES', es1, es_io, es_fd, ES_PRIORITY)
    build_sheet('2호선ES', es2, es_io, es_fd, ES_PRIORITY)
    wb_out.save(str(out_path))


def _detect_repeats(rows, threshold=REPEAT_THRESHOLD):
    """분기 내 동일 (역사, 장애종류) 반복 장애 탐지"""
    counter = Counter(
        (r.get('역사',''), r.get('장애종류',''))
        for r in rows if r.get('역사') and r.get('장애종류')
    )
    result = []
    for (yeoksa, fault_type), cnt in sorted(counter.items(), key=lambda x: -x[1]):
        if cnt >= threshold:
            level = '즉시 점검' if cnt >= 5 else '예방 점검 권고'
            color = 'C0392B' if cnt >= 5 else 'E57C0B'
            result.append({'yeoksa': yeoksa, 'fault_type': fault_type,
                           'count': cnt, 'level': level, 'color': color})
    return result


def extract_stats(xlsx_path):
    wb = load_workbook(str(xlsx_path))

    def load_rows(sn):
        ws = wb[sn]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {ws.cell(1, c).value: ws.cell(r, c).value
                   for c in range(1, ws.max_column + 1)}
            if row.get('순번'): rows.append(row)
        return rows

    def monthly(rows):
        mc = defaultdict(int)
        for r in rows:
            d = r.get('발생일자')
            if isinstance(d, datetime.datetime): mc[f"{d.month}월"] += 1
        return dict(sorted(mc.items(), key=lambda x: int(x[0].replace('월',''))))

    def top_stations(rows, n=6):
        sc = Counter(r.get('역사') for r in rows if r.get('역사'))
        return dict(sc.most_common(n))

    def fault_cnt(rows):
        fc = Counter(r.get('장애종류') for r in rows if r.get('장애종류'))
        return dict(fc.most_common())

    def io_cnt(rows):
        return dict(Counter(r.get('옥내/옥외') for r in rows if r.get('옥내/옥외')))

    def avg_rec(rows):
        times = [r.get('총복구시간') for r in rows if isinstance(r.get('총복구시간'), datetime.time)]
        vals = [t.hour*60+t.minute for t in times]
        return round(sum(vals)/len(vals), 1) if vals else 0

    def make_line(rows, equip):
        return {'total': len(rows), 'monthly': monthly(rows), 'stations': top_stations(rows),
                'faults': fault_cnt(rows), 'indoor': io_cnt(rows).get('옥내형',0),
                'outdoor': io_cnt(rows).get('옥외형',0), 'avgRecovery': avg_rec(rows),
                'equipment': equip}

    el1 = load_rows('1호선EL'); el2 = load_rows('2호선EL')
    es1 = load_rows('1호선ES'); es2 = load_rows('2호선ES')

    all_months = [r.get('발생일자').month
                  for rows in [el1,el2,es1,es2] for r in rows
                  if isinstance(r.get('발생일자'), datetime.datetime)]

    if all_months:
        max_m, min_m = max(all_months), min(all_months)
        quarter = '1분기' if max_m<=3 else '2분기' if max_m<=6 else '3분기' if max_m<=9 else '4분기'
        yr = el1[0]['발생일자'].year if el1 else datetime.date.today().year
        period = f"{min_m}월 ~ {max_m}월"
    else:
        quarter, yr, period = '4분기', 2025, '10월 ~ 12월'

    return {
        'quarter': quarter, 'year': yr, 'period': period,
        'el': {
            'line1': make_line(el1, 124), 'line2': make_line(el2, 84),
            'repeats': _detect_repeats(el1 + el2),   # ← 신규
        },
        'es': {
            'line1': make_line(es1, 194), 'line2': make_line(es2, 194),
            'repeats': _detect_repeats(es1 + es2),   # ← 신규
        },
    }


def compute_yoy(stats_this: dict, stats_last: dict) -> dict:
    """전년 동기 비교 통계 계산"""

    def cmp_monthly(this_l, last_l):
        result = {}
        for m, v in this_l['monthly'].items():
            lv = last_l['monthly'].get(m, 0)
            result[m] = {'this': v, 'last': lv, 'diff': v - lv}
        return result

    def cmp_stations(this_l, last_l, n=6):
        result = {}
        for station, cnt in list(this_l['stations'].items())[:n]:
            lv = last_l['stations'].get(station, 0)
            result[station] = {'this': cnt, 'last': lv, 'diff': cnt - lv}
        return result

    def cmp_faults(this_l1, this_l2, last_l1, last_l2):
        # 1+2호선 합산
        def merge(l1, l2):
            all_k = set(list(l1['faults'])+list(l2['faults']))
            return {k: l1['faults'].get(k,0)+l2['faults'].get(k,0) for k in all_k}
        tf = merge(this_l1, this_l2)
        lf = merge(last_l1, last_l2)
        all_k = set(list(tf)+list(lf))
        result = {k: {'this': tf.get(k,0), 'last': lf.get(k,0),
                       'diff': tf.get(k,0)-lf.get(k,0)} for k in all_k}
        return dict(sorted(result.items(), key=lambda x: -x[1]['this']))

    def make_cmp(key):
        td = stats_this[key]; ld = stats_last[key]
        tl1,tl2 = td['line1'],td['line2']
        ll1,ll2 = ld['line1'],ld['line2']
        tt = tl1['total']+tl2['total']; lt = ll1['total']+ll2['total']
        diff = tt-lt
        return {
            'total_this': tt, 'total_last': lt,
            'total_diff': diff, 'total_pct': round((diff/max(lt,1))*100,1),
            'line1': {'this':tl1['total'],'last':ll1['total'],'diff':tl1['total']-ll1['total'],
                      'pct': round(((tl1['total']-ll1['total'])/max(ll1['total'],1))*100,1)},
            'line2': {'this':tl2['total'],'last':ll2['total'],'diff':tl2['total']-ll2['total'],
                      'pct': round(((tl2['total']-ll2['total'])/max(ll2['total'],1))*100,1)},
            'monthly':  cmp_monthly(tl1, ll1),
            'stations': cmp_stations(tl1, ll1),
            'faults':   cmp_faults(tl1, tl2, ll1, ll2),
            'year_this': stats_this['year'], 'year_last': stats_last['year'],
            'quarter':   stats_this['quarter'],
        }

    return {'el': make_cmp('el'), 'es': make_cmp('es')}
